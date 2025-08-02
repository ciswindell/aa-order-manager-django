"""
Excel Operations Utilities Module

Contains utility classes for Excel worksheet formatting, styling, and writing
operations that are shared across different order processors.
"""

from pathlib import Path
from typing import Dict, List, Optional

import pandas as pd
from openpyxl.styles import Alignment, Font, NamedStyle


class WorksheetStyler:
    """Utility class for applying formatting and styling to Excel worksheets."""

    @classmethod
    def apply_standard_formatting(cls, worksheet, data: pd.DataFrame) -> None:
        """
        Apply standard cell formatting to the entire worksheet.

        This method extracts the standard formatting logic that was duplicated
        between NMSLOOrderProcessor and FederalOrderProcessor. It applies:
        - Calibri 11pt font
        - Left-aligned, top-aligned text
        - Text wrapping enabled
        - Row height auto-adjustment

        Args:
            worksheet: openpyxl worksheet object to format
            data: pandas DataFrame for context (used for determining range)

        Raises:
            ValueError: If worksheet or data parameters are invalid
        """
        if worksheet is None:
            raise ValueError("worksheet cannot be None")
        if not isinstance(data, pd.DataFrame):
            raise ValueError("data must be a pandas DataFrame")

        # Create normal style - extracted from existing processors
        normal_style = NamedStyle(name="normal")
        normal_style.font = Font(name="Calibri", size=11)
        normal_style.alignment = Alignment(
            horizontal="left", vertical="top", wrap_text=True
        )

        # Apply style to all cells (up to row 500 as in original code)
        for row in worksheet.iter_rows(min_row=1, max_row=500):
            for cell in row:
                cell.style = normal_style
            worksheet.row_dimensions[row[0].row].height = None

    @classmethod
    def apply_date_formatting(
        cls, worksheet, data: pd.DataFrame, date_columns: List[str]
    ) -> None:
        """
        Apply "M/D/YYYY" date formatting to specified columns.

        This method extracts the date formatting logic that was duplicated
        between both processors. It applies date formatting to entire columns.

        Args:
            worksheet: openpyxl worksheet object to format
            data: pandas DataFrame containing the data
            date_columns: List of column names to format as dates

        Raises:
            ValueError: If worksheet, data, or date_columns parameters are invalid
        """
        if worksheet is None:
            raise ValueError("worksheet cannot be None")
        if not isinstance(data, pd.DataFrame):
            raise ValueError("data must be a pandas DataFrame")
        if not isinstance(date_columns, list):
            raise ValueError("date_columns must be a list")

        # Apply date formatting to specified columns - extracted from existing processors
        for col_name in date_columns:
            if col_name in data.columns:
                col_idx = data.columns.get_loc(col_name) + 1  # Excel is 1-indexed
                col_letter = chr(ord("A") + col_idx - 1)
                # Format the entire column as date
                for row in range(1, worksheet.max_row + 1):  # Include header
                    cell = worksheet[f"{col_letter}{row}"]
                    cell.number_format = "M/D/YYYY"

    @classmethod
    def apply_column_widths(
        cls, worksheet, data: pd.DataFrame, column_widths: Dict[str, int]
    ) -> None:
        """
        Set column widths based on provided dictionary mapping.

        This method extracts the column width setting logic that was duplicated
        between both processors.

        Args:
            worksheet: openpyxl worksheet object to modify
            data: pandas DataFrame containing the columns
            column_widths: Dictionary mapping column names to width values

        Raises:
            ValueError: If worksheet, data, or column_widths parameters are invalid
        """
        if worksheet is None:
            raise ValueError("worksheet cannot be None")
        if not isinstance(data, pd.DataFrame):
            raise ValueError("data must be a pandas DataFrame")
        if not isinstance(column_widths, dict):
            raise ValueError("column_widths must be a dictionary")

        # Set column widths using column names - extracted from existing processors
        for idx, col in enumerate(data.columns):
            column_letter = chr(ord("A") + idx)
            worksheet.column_dimensions[column_letter].width = column_widths.get(
                col, 12
            )  # Default to 12 if not specified

    @classmethod
    def freeze_header_row(cls, worksheet) -> None:
        """
        Freeze the top row for scrolling.

        This method extracts the freeze panes functionality that was duplicated
        between both processors.

        Args:
            worksheet: openpyxl worksheet object to modify

        Raises:
            ValueError: If worksheet parameter is invalid
        """
        if worksheet is None:
            raise ValueError("worksheet cannot be None")

        # Freeze panes at row 2, column 1 - extracted from existing processors
        worksheet.freeze_panes = worksheet.cell(row=2, column=1)

    @classmethod
    def add_auto_filter(cls, worksheet, data: pd.DataFrame) -> None:
        """
        Add auto-filter to the header row.

        This method extracts the auto-filter functionality that was duplicated
        between both processors.

        Args:
            worksheet: openpyxl worksheet object to modify
            data: pandas DataFrame for determining filter range

        Raises:
            ValueError: If worksheet or data parameters are invalid
        """
        if worksheet is None:
            raise ValueError("worksheet cannot be None")
        if not isinstance(data, pd.DataFrame):
            raise ValueError("data must be a pandas DataFrame")

        # Add auto filter to header row - extracted from existing processors
        worksheet.auto_filter.ref = "A1:{}1".format(chr(ord("A") + data.shape[1] - 1))


class ExcelWriter:
    """Utility class for writing and formatting Excel files."""

    @classmethod
    def save_with_formatting(
        cls, data: pd.DataFrame, output_path: Path, column_widths: Dict[str, int]
    ) -> str:
        """
        Save DataFrame to Excel with complete formatting applied.

        This method extracts the Excel writing and formatting workflow that was
        duplicated between NMSLOOrderProcessor and FederalOrderProcessor. It:
        - Creates ExcelWriter with openpyxl engine
        - Saves DataFrame to "Worksheet" sheet
        - Applies all standard formatting through WorksheetStyler
        - Handles proper file closing

        Args:
            data: pandas DataFrame to save
            output_path: Path object for output file location
            column_widths: Dictionary mapping column names to width values

        Returns:
            str: Path to the created Excel file

        Raises:
            ValueError: If data is not a DataFrame or output_path is invalid
            IOError: If file cannot be written or accessed
        """
        if not isinstance(data, pd.DataFrame):
            raise ValueError("data must be a pandas DataFrame")
        if not isinstance(output_path, Path):
            raise ValueError("output_path must be a Path object")
        if not isinstance(column_widths, dict):
            raise ValueError("column_widths must be a dictionary")

        try:
            # Create ExcelWriter and save data - extracted from existing processors
            writer = pd.ExcelWriter(output_path, engine="openpyxl")
            data.to_excel(writer, index=False, sheet_name="Worksheet")

            # Get worksheet object for formatting
            worksheet = writer.sheets["Worksheet"]

            # Apply all formatting using WorksheetStyler methods
            WorksheetStyler.apply_column_widths(worksheet, data, column_widths)
            WorksheetStyler.apply_standard_formatting(worksheet, data)

            # Apply date formatting to standard date columns - extracted from existing processors
            date_columns = ["Order Date", "Report Start Date"]
            WorksheetStyler.apply_date_formatting(worksheet, data, date_columns)

            # Apply worksheet features
            WorksheetStyler.freeze_header_row(worksheet)
            WorksheetStyler.add_auto_filter(worksheet, data)

            # Close writer properly
            writer.close()

            return str(output_path)

        except PermissionError as e:
            raise IOError(f"Permission denied writing to file: {output_path}") from e
        except FileNotFoundError as e:
            raise IOError(f"Directory not found for file: {output_path}") from e
        except Exception as e:
            raise IOError(f"Error writing Excel file: {str(e)}") from e
