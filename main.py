import pandas as pd
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, NamedStyle

from abc import ABC, abstractmethod


def search_file(lease: str) -> str:
    try:
        number_string = ''.join([x for x in lease if x.isdigit() or x == '0'])
        number = int(number_string)
        search_string = '*' + str(number) + '*'
    except ValueError:
        return 'Error'

    return search_string


def search_tractstar(lease: str) -> str:
    try:
        base_number = ''.join(lease.split(' ')[1:])
        number_string = ''.join([x for x in lease if x.isdigit() or x == '0'])
        alpha_string = ''.join([x for x in base_number if x.isalpha()])
        number = int(number_string)
        if alpha_string:
            search_string = str(number) + '-' + alpha_string
        else:
            search_string = str(number)
    except ValueError:
        return 'Error'

    return search_string


def search_full(lease: str) -> str:
    try:
        search = '*' + lease.replace('-', '*') + '*'
    except ValueError:
        return 'Error'
    return search


def search_partial(lease: str) -> str:
    try:
        search_split = lease.split('-')[:2]
        search = '*' + ('*'.join(search_split)) + '*'
    except ValueError:
        return 'Error'
    return search


class OrderProcessor(ABC):
    def __init__(self, order_form):
        self.order_form = order_form

    @abstractmethod
    def read_order_form(self):
        pass

    @abstractmethod
    def create_order_worksheet(self):
        pass

    @abstractmethod
    def process_data(self):
        pass


class NMStateOrderProcessor(OrderProcessor):
    def read_order_form(self):
        return pd.read_excel(self.order_form)

    def process_data(self) -> pd.DataFrame:
        data = self.read_order_form()

        data['Search_Full'] = data['Lease'].apply(search_full)
        data['Search_Partial'] = data['Lease'].apply(search_partial)
        print(data)

        # Add blank columns for the Old_Format, New_Format, and Tractstar columns
        blank_columns = pd.DataFrame(columns=['Old_Format', 'New_Format', 'Tractstar', 'MI_Index', 'Documents', 'Basecamp'], index=data.index)
        data = pd.concat([data, blank_columns], axis=1)

        return data

    def create_order_worksheet(self):
        data = self.process_data()
        output_path = 'ExampleData/order_worksheet.xlsx'
        writer = pd.ExcelWriter(output_path, engine='openpyxl')
        data.to_excel(writer, index=False, sheet_name='Worksheet')

        worksheet = writer.sheets['Worksheet']
        worksheet.column_dimensions['A'].width = 15
        worksheet.column_dimensions['B'].width = 9
        worksheet.column_dimensions['C'].width = 14
        worksheet.column_dimensions['D'].width = 14
        worksheet.column_dimensions['E'].width = 12
        worksheet.column_dimensions['F'].width = 12
        worksheet.column_dimensions['G'].width = 12
        worksheet.column_dimensions['H'].width = 12
        worksheet.column_dimensions['I'].width = 12
        worksheet.column_dimensions['J'].width = 30

        normal_style = NamedStyle(name="normal")
        normal_style.font = Font(name='Calibri', size=11)
        normal_style.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

        for row in worksheet.iter_rows(min_row=1, max_row=500):
            for cell in row:
                cell.style = normal_style
            worksheet.row_dimensions[row[0].row].height = None

        worksheet.freeze_panes = worksheet.cell(row=2, column=1)
        worksheet.auto_filter.ref = "A1:{}1".format(chr(ord('A') + data.shape[1] - 1))

        writer.close()


class FederalOrderProcessor(OrderProcessor):
    def read_order_form(self):
        return pd.read_excel(self.order_form)

    def process_data(self) -> pd.DataFrame:
        data = self.read_order_form()

        data['Search Files'] = data['Lease'].apply(search_file)
        data['Search Tractstar'] = data['Lease'].apply(search_tractstar)

        blank_columns = pd.DataFrame(columns=['New Format', 'Tractstar', 'Documents', 'Basecamp'], index=data.index)
        data = pd.concat([data, blank_columns], axis=1)

        print(data)

        return data

    def create_order_worksheet(self):
        data = self.process_data()
        output_path = 'ExampleData/order_worksheet.xlsx'
        writer = pd.ExcelWriter(output_path, engine='openpyxl')
        data.to_excel(writer, index=False, sheet_name='Worksheet')

        worksheet = writer.sheets['Worksheet']
        worksheet.column_dimensions['A'].width = 15
        worksheet.column_dimensions['B'].width = 20
        worksheet.column_dimensions['C'].width = 14
        worksheet.column_dimensions['D'].width = 14
        worksheet.column_dimensions['E'].width = 12
        worksheet.column_dimensions['F'].width = 12
        worksheet.column_dimensions['G'].width = 12
        worksheet.column_dimensions['H'].width = 30

        normal_style = NamedStyle(name="normal")
        normal_style.font = Font(name='Calibri', size=11)
        normal_style.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

        for row in worksheet.iter_rows(min_row=1, max_row=500):
            for cell in row:
                cell.style = normal_style
            worksheet.row_dimensions[row[0].row].height = None

        worksheet.freeze_panes = worksheet.cell(row=2, column=1)
        worksheet.auto_filter.ref = "A1:{}1".format(chr(ord('A') + data.shape[1] - 1))

        writer.close()


if __name__ == '__main__':
    # lease_number = 'NMNM 0001375A'
    # parsed = search_tractstar(lease_number)
    # print(lease_number)
    # print(parsed)

    order_form_path = 'ExampleData/order_state.xlsx'
    order_processor = NMStateOrderProcessor(order_form_path)
    order_processor.create_order_worksheet()



