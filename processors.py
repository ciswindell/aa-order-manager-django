from abc import ABC, abstractmethod
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    NamedStyle,
    PatternFill,
    Protection,
    Side,
)


class LeaseNumberParser:
    def __init__(self, lease_number):
        self.lease_number = lease_number

    def search_file(self) -> str:
        try:
            number_string = "".join(
                [x for x in self.lease_number if x.isdigit() or x == "0"]
            )
            number = int(number_string)
            search_string = "*" + str(number) + "*"
        except ValueError:
            return "Error"

        return search_string

    def search_tractstar(self) -> str:
        try:
            base_number = "".join(self.lease_number.split(" ")[1:])
            number_string = "".join(
                [x for x in self.lease_number if x.isdigit() or x == "0"]
            )
            alpha_string = "".join([x for x in base_number if x.isalpha()])
            number = int(number_string)
            if alpha_string:
                search_string = str(number) + "-" + alpha_string
            else:
                search_string = str(number)
        except ValueError:
            return "Error"

        return search_string

    def search_full(self) -> str:
        try:
            search = "*" + self.lease_number.replace("-", "*") + "*"
        except ValueError:
            return "Error"
        return search

    def search_partial(self) -> str:
        try:
            search_split = self.lease_number.split("-")[:2]
            search = "*" + ("*".join(search_split)) + "*"
        except ValueError:
            return "Error"
        return search


class OrderProcessor(ABC):
    def __init__(
        self,
        order_form,
        agency=None,
        order_type=None,
        order_date=None,
        order_number=None,
    ):
        self.order_form = order_form
        self.agency = agency
        self.order_type = order_type
        self.order_date = order_date
        self.order_number = order_number
        self.data = self.read_order_form()

    @abstractmethod
    def read_order_form(self):
        pass

    @abstractmethod
    def create_order_worksheet(self):
        pass

    @abstractmethod
    def process_data(self):
        pass

    @abstractmethod
    def create_folders(self):
        pass


class NMStateOrderProcessor(OrderProcessor):
    def read_order_form(self):
        data = pd.read_excel(self.order_form)

        # Clean Report Start Date column - keep only actual dates, make everything else blank
        if "Report Start Date" in data.columns:

            def clean_date(value):
                if pd.isna(value) or value == "":
                    return None

                # If it's a string, check if it contains only letters (like "Inception")
                if isinstance(value, str):
                    # If it's all letters or contains words like "inception", make it blank
                    if value.isalpha() or "inception" in value.lower():
                        return None
                    # Try to parse date strings
                    try:
                        return pd.to_datetime(value, errors="raise")
                    except:
                        return None

                # If it's a number, try to convert (could be Excel serial date)
                if isinstance(value, (int, float)):
                    try:
                        # Convert Excel serial number to date
                        return pd.to_datetime("1899-12-30") + pd.Timedelta(days=value)
                    except:
                        return None

                # If it's already a datetime, keep it
                if isinstance(value, (pd.Timestamp, datetime)):
                    return value

                return None

            data["Report Start Date"] = data["Report Start Date"].apply(clean_date)

        return data

    def process_data(self) -> pd.DataFrame:
        data = self.data

        # Add new columns if they don't exist, placing them at the beginning
        new_columns = ["Agency", "Order Type", "Order Number", "Order Date"]
        existing_columns = data.columns.tolist()

        # Create empty columns for the new fields
        for col in reversed(
            new_columns
        ):  # Reverse to maintain order when inserting at front
            if col not in existing_columns:
                data.insert(0, col, "")

        # Prefill Agency and Order Type columns based on GUI selections
        if self.agency:
            data["Agency"] = self.agency
        if self.order_type:
            data["Order Type"] = self.order_type
        if self.order_date:
            data["Order Date"] = self.order_date
        if self.order_number:
            data["Order Number"] = self.order_number

        data["Full Search"] = data["Lease"].apply(
            lambda x: LeaseNumberParser(x).search_full()
        )
        data["Partial Search"] = data["Lease"].apply(
            lambda x: LeaseNumberParser(x).search_partial()
        )

        blank_columns = pd.DataFrame(
            columns=[
                "New Format",
                "Tractstar",
                "Old Format",
                "MI Index",
                "Documents",
                "Search Notes",
                "Link",
            ],
            index=data.index,
        )
        data = pd.concat([data, blank_columns], axis=1)

        return data

    def create_order_worksheet(self):
        data = self.process_data()
        date_string = datetime.now().strftime("%Y%m%d")
        base_path = Path(self.order_form).absolute().parent
        file_name = f"{date_string}_nmstate_order_worksheet.xlsx"
        output_path = base_path / file_name
        writer = pd.ExcelWriter(output_path, engine="openpyxl")
        data.to_excel(writer, index=False, sheet_name="Worksheet")

        worksheet = writer.sheets["Worksheet"]

        # Update column widths dictionary to include new columns
        column_widths = {
            "Agency": 15,
            "Order Type": 15,
            "Order Number": 15,
            "Order Date": 15,
            "Lease": 15,
            "Requested Legal": 25,
            "Report Start Date": 20,
            "Full Search": 14,
            "Partial Search": 14,
            "New Format": 12,
            "Tractstar": 12,
            "Old Format": 12,
            "MI Index": 12,
            "Documents": 12,
            "Search Notes": 30,
            "Link": 30,
        }

        # Set column widths using column names
        for idx, col in enumerate(data.columns):
            column_letter = chr(ord("A") + idx)
            worksheet.column_dimensions[column_letter].width = column_widths.get(
                col, 12
            )  # Default to 12 if not specified

        normal_style = NamedStyle(name="normal")
        normal_style.font = Font(name="Calibri", size=11)
        normal_style.alignment = Alignment(
            horizontal="left", vertical="top", wrap_text=True
        )

        for row in worksheet.iter_rows(min_row=1, max_row=500):
            for cell in row:
                cell.style = normal_style
            worksheet.row_dimensions[row[0].row].height = None

        # Apply date formatting to date columns
        date_columns = ["Order Date", "Report Start Date"]
        for col_name in date_columns:
            if col_name in data.columns:
                col_idx = data.columns.get_loc(col_name) + 1  # Excel is 1-indexed
                col_letter = chr(ord("A") + col_idx - 1)
                # Format the entire column as date
                for row in range(1, worksheet.max_row + 1):  # Include header
                    cell = worksheet[f"{col_letter}{row}"]
                    cell.number_format = "M/D/YYYY"

        worksheet.freeze_panes = worksheet.cell(row=2, column=1)
        worksheet.auto_filter.ref = "A1:{}1".format(chr(ord("A") + data.shape[1] - 1))

        writer.close()

    def create_folders(self):
        base_path = Path(self.order_form).absolute().parent
        directories = ["^Document Archive", "^MI Index", "Runsheets"]

        for lease in self.data["Lease"]:
            for directory in directories:
                (base_path / lease / directory).mkdir(exist_ok=True, parents=True)


class FederalOrderProcessor(OrderProcessor):
    def read_order_form(self):
        data = pd.read_excel(self.order_form)

        # Clean Report Start Date column - keep only actual dates, make everything else blank
        if "Report Start Date" in data.columns:

            def clean_date(value):
                if pd.isna(value) or value == "":
                    return None

                # If it's a string, check if it contains only letters (like "Inception")
                if isinstance(value, str):
                    # If it's all letters or contains words like "inception", make it blank
                    if value.isalpha() or "inception" in value.lower():
                        return None
                    # Try to parse date strings
                    try:
                        return pd.to_datetime(value, errors="raise")
                    except:
                        return None

                # If it's a number, try to convert (could be Excel serial date)
                if isinstance(value, (int, float)):
                    try:
                        # Convert Excel serial number to date
                        return pd.to_datetime("1899-12-30") + pd.Timedelta(days=value)
                    except:
                        return None

                # If it's already a datetime, keep it
                if isinstance(value, (pd.Timestamp, datetime)):
                    return value

                return None

            data["Report Start Date"] = data["Report Start Date"].apply(clean_date)

        return data

    def process_data(self) -> pd.DataFrame:
        data = self.data

        # Add new columns if they don't exist, placing them at the beginning
        new_columns = ["Agency", "Order Type", "Order Number", "Order Date"]
        existing_columns = data.columns.tolist()

        # Create empty columns for the new fields
        for col in reversed(
            new_columns
        ):  # Reverse to maintain order when inserting at front
            if col not in existing_columns:
                data.insert(0, col, "")

        # Prefill Agency and Order Type columns based on GUI selections
        if self.agency:
            data["Agency"] = self.agency
        if self.order_type:
            data["Order Type"] = self.order_type
        if self.order_date:
            data["Order Date"] = self.order_date
        if self.order_number:
            data["Order Number"] = self.order_number

        data["Files Search"] = data["Lease"].apply(
            lambda x: LeaseNumberParser(x).search_file()
        )
        data["Tractstar Search"] = data["Lease"].apply(
            lambda x: LeaseNumberParser(x).search_tractstar()
        )

        blank_columns = pd.DataFrame(
            columns=["New Format", "Tractstar", "Documents", "Search Notes", "Link"],
            index=data.index,
        )
        data = pd.concat([data, blank_columns], axis=1)

        return data

    def create_order_worksheet(self):
        data = self.process_data()
        date_string = datetime.now().strftime("%Y%m%d")
        base_path = Path(self.order_form).absolute().parent
        file_name = f"{date_string}_federal_order_worksheet.xlsx"
        output_path = base_path / file_name
        writer = pd.ExcelWriter(output_path, engine="openpyxl")
        data.to_excel(writer, index=False, sheet_name="Worksheet")

        worksheet = writer.sheets["Worksheet"]

        # Define column widths using a dictionary
        column_widths = {
            "Agency": 15,
            "Order Type": 15,
            "Order Number": 15,
            "Order Date": 15,
            "Lease": 15,
            "Requested Legal": 25,
            "Report Start Date": 20,
            "Notes": 30,
            "Files Search": 14,
            "Tractstar Search": 14,
            "New Format": 12,
            "Tractstar": 12,
            "Documents": 12,
            "Search Notes": 30,
            "Link": 30,
        }

        # Set column widths using column names
        for idx, col in enumerate(data.columns):
            column_letter = chr(ord("A") + idx)
            worksheet.column_dimensions[column_letter].width = column_widths.get(
                col, 12
            )  # Default to 12 if not specified

        normal_style = NamedStyle(name="normal")
        normal_style.font = Font(name="Calibri", size=11)
        normal_style.alignment = Alignment(
            horizontal="left", vertical="top", wrap_text=True
        )

        for row in worksheet.iter_rows(min_row=1, max_row=500):
            for cell in row:
                cell.style = normal_style
            worksheet.row_dimensions[row[0].row].height = None

        # Apply date formatting to entire date columns
        date_columns = ["Order Date", "Report Start Date"]
        for col_name in date_columns:
            if col_name in data.columns:
                col_idx = data.columns.get_loc(col_name) + 1  # Excel is 1-indexed
                col_letter = chr(ord("A") + col_idx - 1)
                # Format the entire column as date
                for row in range(1, worksheet.max_row + 1):  # Include header
                    cell = worksheet[f"{col_letter}{row}"]
                    cell.number_format = "M/D/YYYY"

        worksheet.freeze_panes = worksheet.cell(row=2, column=1)
        worksheet.auto_filter.ref = "A1:{}1".format(chr(ord("A") + data.shape[1] - 1))

        writer.close()

    def create_folders(self):
        base_path = Path(self.order_form).absolute().parent
        directories = ["^Document Archive", "Runsheets"]

        for lease in self.data["Lease"]:
            for directory in directories:
                (base_path / lease / directory).mkdir(exist_ok=True, parents=True)


if __name__ == "__main__":
    state_order_form_path = "sample_data/order_state.xlsx"
    federal_order_form_path = "sample_data/order_fed.xlsx"
    order_processor = NMStateOrderProcessor(state_order_form_path)
    # order_processor = FederalOrderProcessor(federal_order_form_path)
    order_processor.create_order_worksheet()
